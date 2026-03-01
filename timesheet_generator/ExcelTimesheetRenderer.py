import os
import sys

from openpyxl import load_workbook

from .cell_style_generators.DayCellStyleGenerator import DayCellStyleGenerator
from .cell_style_generators.DateCellStyleGenerator import DateCellStyleGenerator
from .cell_style_generators.AttendanceCellStyleGenerator import AttendanceCellStyleGenerator

# pyinstaller stores file in a temporary folder on users system during the program execution
# Location of that temporary folder is set in sys._MEIPASS by pyinstaller
# In case that path does not exist we use current directory as base bath
def resource_path(relative_path):
  try:
    base_path = sys._MEIPASS
  except Exception:
    base_path = os.path.abspath(".")

  return os.path.join(base_path, relative_path)

class ExcelTimesheetRenderer:
  def __init__(self, employee_name, month_name, sheet_name, timesheet_data):
    self.employee_name = employee_name
    self.month_name = month_name
    self.sheet_name = sheet_name
    self.timesheet_data = timesheet_data
    self.workbook = self.create_workbook()
    self.worksheet = self.workbook[self.sheet_name]

  def create_workbook(self):
    workbook = load_workbook(resource_path("timesheet_template.xlsx"))
    workbook.active = workbook["template_sheet"]
    workbook["template_sheet"].title = self.sheet_name
    return workbook

  def render(self, file_name):
    self.render_month_name()
    self.render_employee_name()
    self.render_timesheet_data()
    self.save_workbook(file_name)

  def render_month_name(self):
    self.worksheet["A1"] = self.month_name

  def render_employee_name(self):
    self.worksheet["B5"] = self.employee_name

  def render_timesheet_data(self):
    for index, timesheet_date in enumerate(self.timesheet_data):
      self.render_day(self.worksheet.cell(row=3, column=(3 + index)), timesheet_date)
      self.render_date(self.worksheet.cell(row=4, column=(3 + index)), timesheet_date)
      self.render_attendance(self.worksheet.cell(row=5, column=(3 + index)), timesheet_date)

  def render_day(self, cell, timesheet_date):
    day = timesheet_date["day"]
    styles = DayCellStyleGenerator(day).generate_styles()
    self.render_cell(cell, styles)

  def render_date(self, cell, timesheet_date):
    attendance_date = timesheet_date["attendance_date"]
    styles = DateCellStyleGenerator(attendance_date).generate_styles()
    self.render_cell(cell, styles)
  
  def render_attendance(self, cell, timesheet_date):
    attendance = timesheet_date["attendance"]
    styles = AttendanceCellStyleGenerator(attendance).generate_styles()
    self.render_cell(cell, styles)

  def render_cell(self, cell, styles={}):
    cell.value = styles["value"]
    cell.fill = styles["fill"]
    cell.alignment = styles["alignment"]
    cell.border = styles["border"]
    cell.font = styles["font"]
    if "number_format" in styles:
      cell.number_format = styles["number_format"]

  def save_workbook(self, filename):
    self.workbook.save(filename)
    self.workbook.close()